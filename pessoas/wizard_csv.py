import csv
import datetime
import difflib
import io
import re

from openpyxl import load_workbook

from .models import Profissao
from .validators import normalizar_data_nascimento

CAMPOS_CSV = {
    "nome completo": "nome",
    "nome": "nome",
    "cpf": "cpf",
    "data de nascimento": "data_nascimento",
    "data nascimento": "data_nascimento",
    "nascimento": "data_nascimento",
    "genero": "genero",
    "gênero": "genero",
    "raca": "raca",
    "raça": "raca",
    "raca/cor": "raca",
    "raça/cor": "raca",
    "e-mail": "email",
    "email": "email",
    "telefone": "telefone",
    "cidade": "cidade",
    "uf": "uf",
    "cep": "cep",
    "bairro": "bairro",
    "regiao": "regiao",
    "região": "regiao",
    "escolaridade": "escolaridade",
    "profissao": "profissao",
    "profissão": "profissao",
    "especialidade": "especialidade",
    "ocupacao": "ocupacao",
    "ocupação": "ocupacao",
    "estado civil": "estado_civil",
    "estado_civil": "estado_civil",
    "renda individual": "renda_individual",
    "renda_individual": "renda_individual",
    "faixa de renda": "renda_individual",
    "renda familiar": "renda_familiar",
    "renda_familiar": "renda_familiar",
}

GENERO_MAP = {
    "mulher cisgenero": "MULHER_CIS",
    "mulher cisgênero": "MULHER_CIS",
    "mulher cis": "MULHER_CIS",
    "feminino": "MULHER_CIS",
    "f": "MULHER_CIS",
    "homem cisgenero": "HOMEM_CIS",
    "homem cisgênero": "HOMEM_CIS",
    "homem cis": "HOMEM_CIS",
    "masculino": "HOMEM_CIS",
    "m": "HOMEM_CIS",
    "mulher transgenero": "MULHER_TRANS",
    "mulher transgênero": "MULHER_TRANS",
    "mulher trans": "MULHER_TRANS",
    "homem transgenero": "HOMEM_TRANS",
    "homem transgênero": "HOMEM_TRANS",
    "homem trans": "HOMEM_TRANS",
    "pessoa nao binaria": "NAO_BINARIA",
    "pessoa não binária": "NAO_BINARIA",
    "nao binaria": "NAO_BINARIA",
    "não binária": "NAO_BINARIA",
    "outra identidade de genero": "OUTRA",
    "outra identidade de gênero": "OUTRA",
    "outro": "OUTRA",
    "outra": "OUTRA",
    "prefiro nao responder": "NAO_RESPONDE",
    "prefiro não responder": "NAO_RESPONDE",
    "nao informa": "NAO_RESPONDE",
    "não informa": "NAO_RESPONDE",
}
RACA_MAP = {
    "branca": "BRANCA",
    "branco": "BRANCA",
    "branco(a)": "BRANCA",
    "preta": "PRETA",
    "preto": "PRETA",
    "preto(a)": "PRETA",
    "parda": "PARDA",
    "pardo": "PARDA",
    "pardo(a)": "PARDA",
    "amarela": "AMARELA",
    "amarelo": "AMARELA",
    "amarelo(a)": "AMARELA",
    "indigena": "INDIGENA",
    "indígena": "INDIGENA",
    "indigena(a)": "INDIGENA",
    "indígena(a)": "INDIGENA",
}
ESTADO_CIVIL_MAP = {
    "solteiro": "SOLTEIRO",
    "solteiro(a)": "SOLTEIRO",
    "casado": "CASADO",
    "casado(a)": "CASADO",
    "uniao estavel": "UNIAO_ESTAVEL",
    "união estável": "UNIAO_ESTAVEL",
    "separado": "SEPARADO",
    "separado(a)": "SEPARADO",
    "divorciado": "DIVORCIADO",
    "divorciado(a)": "DIVORCIADO",
    "viuvo": "VIUVO",
    "viúvo": "VIUVO",
    "viuvo(a)": "VIUVO",
    "viúvo(a)": "VIUVO",
}
OCUPACAO_MAP = {
    "ocupacao remunerada": "OCUPACAO_REMUNERADA",
    "ocupação remunerada": "OCUPACAO_REMUNERADA",
    "estudante e ocupacao remunerada": "ESTUDANTE_E_OCUPACAO",
    "estudante e ocupação remunerada": "ESTUDANTE_E_OCUPACAO",
    "estudante": "ESTUDANTE",
    "desempregado": "DESEMPREGADO",
    "desempregado(a)": "DESEMPREGADO",
    "aposentado": "APOSENTADO",
    "aposentado(a)": "APOSENTADO",
    "atividades do lar": "ATIVIDADES_DO_LAR",
}
REGIAO_MAP = {
    "norte": "NORTE",
    "nordeste": "NORDESTE",
    "centro-oeste": "CENTRO_OESTE",
    "centro oeste": "CENTRO_OESTE",
    "sudeste": "SUDESTE",
    "sul": "SUL",
}
ESCOLARIDADE_MAP = {
    "medio incompleto": "MEDIO_INCOMPLETO",
    "médio incompleto": "MEDIO_INCOMPLETO",
    "ensino medio incompleto": "MEDIO_INCOMPLETO",
    "ensino médio incompleto": "MEDIO_INCOMPLETO",
    "medio completo": "MEDIO_COMPLETO",
    "médio completo": "MEDIO_COMPLETO",
    "ensino medio completo": "MEDIO_COMPLETO",
    "ensino médio completo": "MEDIO_COMPLETO",
    "medio": "MEDIO_COMPLETO",
    "médio": "MEDIO_COMPLETO",
    "superior incompleto": "SUPERIOR_INCOMPLETO",
    "ensino superior incompleto": "SUPERIOR_INCOMPLETO",
    "superior completo": "SUPERIOR_COMPLETO",
    "ensino superior completo": "SUPERIOR_COMPLETO",
    "superior": "SUPERIOR_COMPLETO",
    "pos": "POS_GRADUACAO",
    "pós": "POS_GRADUACAO",
    "pos graduacao": "POS_GRADUACAO",
    "pós graduação": "POS_GRADUACAO",
    "pos-graduacao": "POS_GRADUACAO",
    "pós-graduação": "POS_GRADUACAO",
    "mestrado": "MESTRADO",
    "doutorado": "DOUTORADO",
}
# Mesmos códigos A-E servem pra renda individual e familiar — o rótulo (o
# valor em R$/salários mínimos) é que muda por campo, o código de classe é
# igual nos dois.
RENDA_MAP = {
    "a": "A",
    "classe a": "A",
    "b": "B",
    "classe b": "B",
    "c": "C",
    "classe c": "C",
    "d": "D",
    "classe d": "D",
    "e": "E",
    "classe e": "E",
}

# Fallback pra quando a faixa de renda vem descrita em R$ solto (ex.: "Acima
# de R$ 10788.56", "Entre R$ 5721.73 e R$ 10788.56") em vez de letra ou
# "classe X" — não bate com `RENDA_MAP`, então sem isso a faixa inteira era
# perdida (virava "cadastro incompleto" mesmo com o dado presente na
# planilha). Individual e familiar usam escalas DIFERENTES pro mesmo código
# de classe (ver `Participante.FaixaRendaIndividual`/`FaixaRendaFamiliar`) —
# individual é R$ direto, familiar é múltiplo de salário mínimo — por isso
# cada campo tem sua própria tabela de limiares, nunca a mesma.
_FAIXAS_RENDA_INDIVIDUAL_POR_VALOR = [(9738, "A"), (4869, "B"), (1883, "C"), (974, "D")]

# Salário mínimo vigente usado só pra converter a faixa familiar (múltiplos
# de salário mínimo) em limiares de R$ — não é lido de lugar nenhum porque o
# sistema não guarda esse valor em nenhum outro ponto; ajustar aqui quando o
# valor real mudar.
SALARIO_MINIMO = 1518
_FAIXAS_RENDA_FAMILIAR_POR_VALOR = [
    (20 * SALARIO_MINIMO, "A"),
    (10 * SALARIO_MINIMO, "B"),
    (4 * SALARIO_MINIMO, "C"),
    (2 * SALARIO_MINIMO, "D"),
]


def _numeros_do_texto(texto):
    brutos = re.findall(r"\d[\d.,]*\d|\d", texto)
    valores = []
    for bruto in brutos:
        try:
            valores.append(float(bruto.replace(",", "")))
        except ValueError:
            continue
    return valores


def _mapear_renda_por_valor(texto, faixas):
    """"Acima de R$ 10788.56" → código conforme `faixas`; "Entre R$ 5721.73 e
    R$ 10788.56" usa o limite de baixo da faixa — é ele que decide em qual
    balde ela cai: uma faixa "entre 5721 e 10788" é claramente o balde
    abaixo de "acima de 10788". Devolve "" se não achar nenhum número no
    texto."""
    valores = _numeros_do_texto(texto)
    if not valores:
        return ""
    valor = min(valores)
    for limite, codigo in faixas:
        if valor >= limite:
            return codigo
    return "E"

CABECALHO_MODELO = [
    "Nome completo",
    "CPF",
    "Data de nascimento",
    "Gênero",
    "Raça/cor",
    "E-mail",
    "Telefone",
    "Cidade",
    "UF",
    "Região",
    "CEP",
    "Bairro",
    "Escolaridade",
    "Profissão",
    "Especialidade",
    "Ocupação",
    "Estado civil",
    "Renda individual",
    "Renda familiar",
]

LINHA_EXEMPLO = [
    "Maria da Silva",
    "111.444.777-35",
    "1990-05-10",
    "Mulher cisgênero",
    "Branca",
    "maria@example.com",
    "11999998888",
    "São Paulo",
    "SP",
    "Sudeste",
    "01000-000",
    "Centro",
    "Superior Completo",
    "Designer",
    "UX/UI",
    "Ocupação remunerada",
    "Solteiro(a)",
    "Classe B",
    "Classe B",
]


def _mapear(valor, mapa):
    return mapa.get((valor or "").strip().lower(), "")


def _mapa_profissoes():
    """Nome (minúsculo) → PK, montado com uma única consulta e reaproveitado
    pra todas as linhas do arquivo — evita uma query por célula."""
    return {p.nome.lower(): str(p.pk) for p in Profissao.objects.all()}


# Abaixo de que nível de parecença (0-1, escala do `difflib`) uma profissão
# da planilha deixa de "bater" com uma já cadastrada — evita casar coisas
# muito diferentes só porque compartilham algumas letras (ex.: um texto
# muito curto tipo "TI" bateria com qualquer profissão que tenha essas duas
# letras). 0.6 cobre variação de gênero gramatical ("Advogada" → "Advogado
# (a)"), acento faltando ("Medico" → "Médico(a)") e abreviação parcial
# ("Enfermeira" → "Enfermeiro(a)") sem casar profissões sem relação nenhuma.
_CORTE_PROFISSAO_PARECIDA = 0.6


def _profissao_por_nome_ou_criar(texto, mapa_profissoes):
    """Resolve o texto livre de profissão da planilha pro PK de uma
    `Profissao` cadastrada: bate exato primeiro, senão pega a mais parecida
    pelo `difflib` (de→para automático, sem precisar de um mapa manual pra
    cada uma das profissões — a lista cresce/muda pelo cadastro, não dava
    pra manter isso hardcoded). Não achando nada parecido o bastante,
    cadastra o texto como profissão nova em vez de descartar o dado —
    `mapa_profissoes` é atualizado na hora, então outra linha do mesmo lote
    com o mesmo texto reaproveita a profissão recém-criada em vez de tentar
    duplicar (o campo `nome` é `unique`)."""
    texto_norm = (texto or "").strip()
    if not texto_norm:
        return ""
    chave = texto_norm.lower()
    pk = mapa_profissoes.get(chave)
    if pk:
        return pk
    proximas = difflib.get_close_matches(chave, mapa_profissoes.keys(), n=1, cutoff=_CORTE_PROFISSAO_PARECIDA)
    if proximas:
        return mapa_profissoes[proximas[0]]
    profissao, _criada = Profissao.objects.get_or_create(nome__iexact=texto_norm, defaults={"nome": texto_norm})
    mapa_profissoes[chave] = str(profissao.pk)
    return str(profissao.pk)


_MAPAS_POR_CAMPO = {
    "genero": GENERO_MAP,
    "raca": RACA_MAP,
    "estado_civil": ESTADO_CIVIL_MAP,
    "ocupacao": OCUPACAO_MAP,
    "regiao": REGIAO_MAP,
    "escolaridade": ESCOLARIDADE_MAP,
    "renda_individual": RENDA_MAP,
    "renda_familiar": RENDA_MAP,
}


def _normalizar_campo(campo, valor, mapa_profissoes=None):
    if campo == "profissao":
        if mapa_profissoes is None:
            mapa_profissoes = _mapa_profissoes()
        return _profissao_por_nome_ou_criar(valor, mapa_profissoes)
    if campo == "data_nascimento":
        return normalizar_data_nascimento(valor)
    if campo in ("renda_individual", "renda_familiar"):
        codigo = _mapear(valor, RENDA_MAP)
        if codigo:
            return codigo
        faixas = (
            _FAIXAS_RENDA_INDIVIDUAL_POR_VALOR
            if campo == "renda_individual"
            else _FAIXAS_RENDA_FAMILIAR_POR_VALOR
        )
        return _mapear_renda_por_valor(valor, faixas)
    mapa = _MAPAS_POR_CAMPO.get(campo)
    if mapa is not None:
        return _mapear(valor, mapa)
    return valor


def variaveis_do_formulario(formulario):
    """Lista ordenada de `FormularioVariavel` de um formulário — mesma
    consulta que `formularios/respostas.py::construir_form_resposta` usa,
    reaproveitada aqui pra montar o cabeçalho/exemplo do modelo de planilha
    e pra saber quais colunas extras esperar na hora de ler o arquivo."""
    return list(
        formulario.formulario_variaveis.select_related("variavel__tipo_resposta")
        .prefetch_related("variavel__opcoes")
        .order_by("ordem")
    )


def variaveis_dos_formularios(formularios):
    """Mesma coisa que `variaveis_do_formulario()`, só que concatenando as
    variáveis de **todos** os formulários do perfil, na ordem
    formulário→variável-dentro-do-formulário, sem repetir uma `Variavel`
    que apareça em mais de um formulário do mesmo perfil (a `chave` da
    variável já é única globalmente, então dedup por chave é suficiente)."""
    vistas = set()
    resultado = []
    for formulario in formularios or []:
        for fv in variaveis_do_formulario(formulario):
            if fv.variavel.chave in vistas:
                continue
            vistas.add(fv.variavel.chave)
            resultado.append(fv)
    return resultado


def _exemplo_variavel(variavel):
    """Valor de exemplo pra célula dessa variável no modelo de planilha —
    mesma lógica de tipo que `formularios/respostas.py::_campo_para_variavel`,
    só que devolvendo um texto pronto pra célula em vez de um widget."""
    tipo = variavel.tipo_resposta.codigo
    if tipo in ("select", "radio"):
        opcao = variavel.opcoes.first()
        return opcao.valor if opcao else ""
    if tipo == "multipla_escolha":
        opcoes = list(variavel.opcoes.all()[:2])
        return ", ".join(o.valor for o in opcoes) if opcoes else ""
    if tipo == "booleano":
        return "Sim"
    if tipo == "data":
        return "2000-01-01"
    if tipo in ("inteiro", "decimal"):
        return "0"
    return "Exemplo"


def _normalizar_valor_dinamico(variavel, valor):
    """Normaliza o texto de uma célula pro formato que
    `construir_form_resposta()` espera validar — só mexe nos tipos que têm
    um vocabulário fixo (booleano/opções); texto, número e data passam como
    a pessoa digitou (o próprio `Field` do formulário dinâmico valida)."""
    tipo = variavel.tipo_resposta.codigo
    valor = (valor or "").strip()
    if tipo == "booleano":
        chave = valor.lower()
        if chave in ("sim", "s", "yes", "verdadeiro", "true", "1"):
            return "sim"
        if chave in ("nao", "não", "n", "no", "falso", "false", "0"):
            return "nao"
        return valor
    if tipo in ("select", "radio", "multipla_escolha"):
        opcoes = {o.valor.strip().lower(): o.valor for o in variavel.opcoes.all()}
        if tipo == "multipla_escolha":
            partes = [p.strip() for p in valor.split(",") if p.strip()]
            return [opcoes.get(p.lower(), p) for p in partes]
        return opcoes.get(valor.lower(), valor)
    return valor


def _normalizar_cabecalho(texto):
    """"Nome Completo *" (o `*` marca obrigatório no modelo baixado) vira
    "nome completo" — sem isso a coluna nunca bate com a variável na hora
    de ler o arquivo de volta."""
    texto = str(texto or "").strip().lower()
    if texto.endswith("*"):
        texto = texto[:-1].strip()
    return texto


def _mapa_variaveis(formularios):
    """Nome da variável (minúsculo) → Variavel — usado pra casar o
    cabeçalho da planilha com a variável certa de algum dos formulários do
    perfil (todos misturados: o cabeçalho da planilha não distingue de qual
    formulário veio cada pergunta)."""
    if not formularios:
        return {}
    return {
        fv.variavel.nome.strip().lower(): fv.variavel
        for fv in variaveis_dos_formularios(formularios)
    }


def _texto_celula(valor):
    """Normaliza o valor de uma célula do Excel pra texto. Datas viram
    'AAAA-MM-DD' (o `openpyxl` já devolve `datetime.date` pra células
    formatadas como data) e números inteiros guardados como float (ex.: CPF
    digitado só com dígitos, que o Excel trata como número) voltam a virar
    string sem casas decimais, senão "11144477735" vira "11144477735.0"."""
    if valor is None:
        return ""
    if isinstance(valor, (datetime.date, datetime.datetime)):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def ler_csv(arquivo, formularios=None):
    """Lê um arquivo CSV enviado via upload e devolve uma lista de dicionários com as
    chaves já normalizadas para os nomes de campo do modelo Participante. Datas devem
    estar no formato AAAA-MM-DD; valores de gênero/escolaridade/faixa de renda aceitam
    variações comuns em português (ex.: "Superior", "Médio", "Classes A/B"). Se
    `formularios` for passado (os formulários do perfil escolhido no wizard), colunas
    que baterem com o nome de uma variável de algum desses formulários também entram
    no resultado, com a chave da variável — prontas pra `construir_form_resposta()`."""
    bruto = arquivo.read()
    for codificacao in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = bruto.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue
    else:
        texto = bruto.decode("utf-8", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(texto[:2000], delimiters=",;")
    except csv.Error:
        dialect = csv.excel

    leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)
    mapa_profissoes = _mapa_profissoes()
    mapa_variaveis = _mapa_variaveis(formularios)
    linhas = []
    for linha_bruta in leitor:
        dados = {}
        for chave_coluna, valor in linha_bruta.items():
            chave_coluna = _normalizar_cabecalho(chave_coluna)
            valor = (valor or "").strip()
            campo = CAMPOS_CSV.get(chave_coluna)
            if campo:
                dados[campo] = _normalizar_campo(campo, valor, mapa_profissoes)
                continue
            variavel = mapa_variaveis.get(chave_coluna)
            if variavel:
                dados[variavel.chave] = _normalizar_valor_dinamico(variavel, valor)
        if any(dados.values()):
            linhas.append(dados)
    return linhas


def ler_xlsx(arquivo, formularios=None):
    """Lê a primeira planilha de um arquivo .xlsx enviado via upload — mesma
    saída de `ler_csv()` (lista de dicionários com os campos já normalizados,
    incluindo as colunas dos `formularios` do perfil quando houver), só que
    lendo célula a célula em vez de linha de texto."""
    pasta = load_workbook(arquivo, data_only=True, read_only=True)
    aba = pasta.active
    linhas_brutas = aba.iter_rows(values_only=True)
    try:
        cabecalho_bruto = next(linhas_brutas)
    except StopIteration:
        return []
    cabecalho = [_normalizar_cabecalho(c) for c in cabecalho_bruto]

    mapa_profissoes = _mapa_profissoes()
    mapa_variaveis = _mapa_variaveis(formularios)
    linhas = []
    for linha_bruta in linhas_brutas:
        dados = {}
        for chave_coluna, valor in zip(cabecalho, linha_bruta):
            campo = CAMPOS_CSV.get(chave_coluna)
            if campo:
                dados[campo] = _normalizar_campo(campo, _texto_celula(valor), mapa_profissoes)
                continue
            variavel = mapa_variaveis.get(chave_coluna)
            if variavel:
                dados[variavel.chave] = _normalizar_valor_dinamico(variavel, _texto_celula(valor))
        if any(dados.values()):
            linhas.append(dados)
    return linhas


def ler_planilha(arquivo, formularios=None):
    """Ponto de entrada único do wizard de importação: aceita tanto .csv
    quanto .xlsx e despacha pro leitor certo com base na extensão do
    arquivo enviado. `formularios` (opcional) são os formulários do perfil
    escolhido no wizard — quando presentes, as respostas deles também são
    lidas da planilha."""
    nome = (arquivo.name or "").lower()
    if nome.endswith(".xlsx"):
        return ler_xlsx(arquivo, formularios)
    return ler_csv(arquivo, formularios)
