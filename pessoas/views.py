import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CAMPOS_PII_REVELAVEIS = ("cpf", "telefone", "email")

from accounts.models import Usuario
from accounts.permissions import requer_permissao
from auditoria.models import RegistroAcesso, registrar, registrar_download, verificar_limite_download
from formularios.models import RespostaFormulario
from formularios.respostas import construir_form_resposta
from participacoes.models import Participacao
from projetos.models import Perfil, Projeto
from termos.models import AceiteTermo, Termo, registrar_aceite

from . import exportacao
from .forms import (
    CadastroPublicoForm,
    EscolherProjetoWizardForm,
    ParticipanteForm,
    ParticipanteWizardForm,
    RenovarTermoForm,
    UploadCSVForm,
    participante_wizard_formset,
)
from .links import gerar_token_renovacao_termo, ler_token_captacao, ler_token_renovacao_termo
from .matching import (
    aplicar_dados_legado,
    capturar_valores_atuais,
    encontrar_participante_existente,
    preparar_linha_legado,
    restaurar_campos_vazios,
)
from .models import Participante
from .validators import normalizar_cpf
from .wizard_csv import (
    CABECALHO_MODELO,
    LINHA_EXEMPLO,
    _exemplo_variavel,
    ler_planilha,
    variaveis_dos_formularios,
)

WIZ_SESSION_KEY = "wizard_importacao"
ITENS_POR_PAGINA = 10


def _versao_lgpd_vigente():
    # order_by explícito: se alguém cadastrar mais de um documento do tipo
    # "Consentimento" por engano, sempre usamos o mais antigo (o original),
    # em vez de depender da ordem não garantida do banco.
    termo = Termo.objects.filter(tipo=Termo.Tipo.CONSENTIMENTO).order_by("id").first()
    return termo.versao_vigente if termo else None


def _link_renovacao_termo(request, participante, termo):
    token = gerar_token_renovacao_termo(participante.id, termo.id)
    return request.build_absolute_uri(reverse("pessoas:renovar_termo", args=[token]))


def _termos_pendentes_renovacao(request, participante, aceites_termos):
    """Documentos que a pessoa já aceitou antes, mas cuja versão vigente
    mudou desde então — um link de renovação por documento pendente, pra
    reenviar pra ela ler e aceitar a versão atual. Não cobre documento nunca
    aceito (aqui é só sobre atualizar consentimento já dado, não pedir um
    novo do zero).

    `aceites_termos` já vem carregado (ordenado por `-aceito_em`, o padrão
    do model) — usa isso em vez de fazer outra consulta: o primeiro aceite
    de cada `termo_id` encontrado na lista já é o mais recente."""
    vistos = set()
    pendentes = []
    for aceite in aceites_termos:
        termo_id = aceite.versao.termo_id
        if termo_id in vistos:
            continue
        vistos.add(termo_id)
        termo = aceite.versao.termo
        vigente = termo.versao_vigente
        if vigente and vigente.id != aceite.versao_id:
            pendentes.append(
                {
                    "termo": termo,
                    "versao_aceita": aceite.versao,
                    "versao_vigente": vigente,
                    "link": _link_renovacao_termo(request, participante, termo),
                }
            )
    return pendentes


def _participantes_filtrados(request):
    q = request.GET.get("q", "").strip()
    situacao = request.GET.get("situacao", "").strip()
    renda_individual = request.GET.get("renda_individual", "").strip()
    uf = request.GET.get("uf", "").strip()
    incompleto = request.GET.get("incompleto", "").strip()

    participantes = Participante.objects.select_related("consentimento_versao")
    if q:
        participantes = participantes.filter(
            Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(codigo__icontains=q)
        )
    if situacao:
        participantes = participantes.filter(situacao=situacao)
    if renda_individual:
        participantes = participantes.filter(renda_individual=renda_individual)
    if uf:
        participantes = participantes.filter(uf=uf)
    if incompleto == "1":
        participantes = participantes.filter(cadastro_incompleto=True)

    filtros = {
        "q": q,
        "situacao": situacao,
        "situacao_label": dict(Participante.Situacao.choices).get(situacao, situacao),
        "renda_individual": renda_individual,
        "renda_individual_label": dict(Participante.FaixaRendaIndividual.choices).get(
            renda_individual, renda_individual
        ),
        "uf": uf,
        "incompleto": incompleto,
    }
    return participantes, filtros


@login_required
@requer_permissao("participantes.ver")
@ensure_csrf_cookie
def lista(request):
    participantes, filtros = _participantes_filtrados(request)
    paginator = Paginator(participantes, ITENS_POR_PAGINA)
    page_obj = paginator.get_page(request.GET.get("page"))
    pode_revelar = request.user.tem_permissao("participantes.revelar_pii")
    pode_exportar = request.user.tem_permissao("participantes.exportar")
    return render(
        request,
        "pessoas/lista.html",
        {
            "participantes": page_obj,
            "page_obj": page_obj,
            "q": filtros["q"],
            "filtro_situacao": filtros["situacao"],
            "filtro_renda_individual": filtros["renda_individual"],
            "filtro_uf": filtros["uf"],
            "filtro_incompleto": filtros["incompleto"],
            "situacoes": Participante.Situacao.choices,
            "faixas_renda": Participante.FaixaRendaIndividual.choices,
            "ufs": Participante.UF.choices,
            "pode_revelar": pode_revelar,
            "pode_exportar": pode_exportar,
        },
    )


@login_required
@requer_permissao("participantes.exportar")
def exportar(request, formato):
    if formato not in ("pdf", "xlsx"):
        raise Http404

    pode_baixar, libera_em = verificar_limite_download(request.user, "pessoas", horas=12)
    if not pode_baixar:
        messages.error(
            request,
            "Por segurança, exportações da base de Pessoas são limitadas a 1 a cada 12 horas. "
            f"Você poderá baixar de novo às {timezone.localtime(libera_em).strftime('%d/%m/%Y %H:%M')}.",
        )
        return redirect("pessoas:lista")

    participantes, filtros = _participantes_filtrados(request)
    registrar_download(request.user, "pessoas")
    if formato == "xlsx":
        return exportacao.gerar_xlsx(participantes, filtros)
    return exportacao.gerar_pdf(participantes, filtros)


@login_required
@requer_permissao("participantes.ver")
@ensure_csrf_cookie
def detalhe(request, pk):
    participante = get_object_or_404(
        Participante.objects.select_related("consentimento_versao"), pk=pk
    )
    pode_revelar = request.user.tem_permissao("participantes.revelar_pii")
    pode_pagamento = request.user.tem_permissao("pagamento.ver")
    pode_editar = request.user.tem_permissao("participantes.gerenciar")
    pode_excluir = request.user.tem_permissao("participantes.excluir")
    aceites_termos = list(participante.aceites_termos.select_related("versao__termo", "registrado_por"))
    termos_pendentes = _termos_pendentes_renovacao(request, participante, aceites_termos)
    return render(
        request,
        "pessoas/detalhe.html",
        {
            "participante": participante,
            "pode_revelar": pode_revelar,
            "pode_pagamento": pode_pagamento,
            "pode_editar": pode_editar,
            "pode_excluir": pode_excluir,
            "aceites_termos": aceites_termos,
            "termos_pendentes": termos_pendentes,
        },
    )


@login_required
@requer_permissao("participantes.revelar_pii")
@require_POST
def revelar_campo(request, pk, campo):
    if campo not in CAMPOS_PII_REVELAVEIS:
        raise Http404
    participante = get_object_or_404(Participante, pk=pk)
    valor = getattr(participante, campo)
    registrar(
        request.user,
        participante.codigo,
        RegistroAcesso.Acao.VISUALIZACAO,
        f"Campo '{campo}' revelado sem máscara",
    )
    return JsonResponse({"valor": valor})


@login_required
@requer_permissao("participantes.excluir")
def excluir(request, pk):
    participante = get_object_or_404(Participante, pk=pk)
    if request.method == "POST":
        nome = participante.nome
        codigo = participante.codigo
        participante.delete()
        registrar(request.user, codigo, RegistroAcesso.Acao.ALTERACAO, f"Participante {nome} excluído")
        messages.success(request, f"{nome} foi excluído(a) do Banco de Pessoas.")
        return redirect("pessoas:lista")
    return render(
        request,
        "pessoas/excluir.html",
        {"participante": participante, "total_participacoes": participante.participacoes.count()},
    )


@login_required
@requer_permissao("participantes.gerenciar")
def novo(request):
    pode_pagamento = request.user.tem_permissao("pagamento.ver")
    if request.method == "POST":
        # Casa por CPF/e-mail/telefone antes de validar o form — se achar,
        # o form é ligado a essa instância (em vez de criar uma nova) e a
        # submissão vira uma atualização dos dados dela.
        existente = encontrar_participante_existente(
            request.POST.get("cpf", ""), request.POST.get("email", ""), request.POST.get("telefone", "")
        )
        valores_originais = capturar_valores_atuais(existente) if existente else None
        form = ParticipanteForm(request.POST, instance=existente, pode_ver_pagamento=pode_pagamento)
        if form.is_valid():
            participante = form.save(commit=False)
            if existente:
                # Uma submissão que vem com campo vazio não apaga o que já
                # tinha sido salvo — só atualiza o que veio preenchido.
                restaurar_campos_vazios(participante, form.cleaned_data, valores_originais)
            else:
                participante.criado_por = request.user
            novo_consentimento = participante.consentimento_lgpd and not participante.consentimento_versao_id
            if novo_consentimento:
                participante.consentimento_versao = _versao_lgpd_vigente()
            participante.save()
            if novo_consentimento and participante.consentimento_versao_id:
                registrar_aceite(
                    participante, participante.consentimento_versao,
                    origem=AceiteTermo.Origem.STAFF, request=request, registrado_por=request.user,
                )
            if existente:
                messages.success(
                    request, f"{participante.codigo} já estava cadastrado(a) — dados atualizados."
                )
            else:
                messages.success(request, f"Participante {participante.codigo} cadastrado com sucesso.")
            return redirect("pessoas:detalhe", pk=participante.pk)
    else:
        form = ParticipanteForm(pode_ver_pagamento=pode_pagamento)
    return render(request, "pessoas/form.html", {"form": form, "titulo": "Novo participante"})


@login_required
@requer_permissao("participantes.gerenciar")
def editar(request, pk):
    participante = get_object_or_404(Participante, pk=pk)
    pode_pagamento = request.user.tem_permissao("pagamento.ver")
    if request.method == "POST":
        form = ParticipanteForm(request.POST, instance=participante, pode_ver_pagamento=pode_pagamento)
        if form.is_valid():
            participante = form.save(commit=False)
            novo_consentimento = participante.consentimento_lgpd and not participante.consentimento_versao_id
            if novo_consentimento:
                participante.consentimento_versao = _versao_lgpd_vigente()
            participante.save()
            if novo_consentimento and participante.consentimento_versao_id:
                registrar_aceite(
                    participante, participante.consentimento_versao,
                    origem=AceiteTermo.Origem.STAFF, request=request, registrado_por=request.user,
                )
            messages.success(request, "Dados do participante atualizados.")
            return redirect("pessoas:detalhe", pk=participante.pk)
    else:
        form = ParticipanteForm(instance=participante, pode_ver_pagamento=pode_pagamento)
    return render(
        request,
        "pessoas/form.html",
        {"form": form, "titulo": f"Editar {participante.nome}"},
    )


# =========================================================================
# Wizard de cadastro em lote (novos participantes)
# =========================================================================

WIZ_STEPS = ["Banco de dados", "Novos participantes", "Dados", "Revisão"]


def _validar_linha_csv(dados, formularios=None, legado=False):
    dados = dict(dados)
    dados.setdefault("situacao", Participante.Situacao.PENDENTE)

    if legado:
        # Lote legado não passa pelo ParticipanteWizardForm — só recusa o
        # que não dá nem pra identificar (sem nome, ou sem nenhuma das três
        # chaves). O resto é aceito como veio, com placeholder nos campos
        # obrigatórios que faltarem (ver `preparar_linha_legado`).
        dados_prontos, campos_incompletos, erro = preparar_linha_legado(dados)
        if erro:
            return {
                "dados": dados,
                "valido": False,
                "erros": {"Linha": [erro]},
                "consentimento": False,
                "existente_pk": None,
                "existente_codigo": "",
                "legado": True,
            }
        existente = encontrar_participante_existente(
            dados_prontos.get("cpf") or "", dados_prontos.get("email", ""), dados_prontos.get("telefone") or ""
        )
        return {
            "dados": dados_prontos,
            "valido": True,
            "erros": None,
            "consentimento": False,
            "existente_pk": existente.pk if existente else None,
            "existente_codigo": existente.codigo if existente else "",
            "legado": True,
            "cadastro_incompleto": bool(campos_incompletos),
        }

    existente = encontrar_participante_existente(
        dados.get("cpf", ""), dados.get("email", ""), dados.get("telefone", "")
    )
    form = ParticipanteWizardForm(data=dados, instance=existente)
    valido = form.is_valid()
    erros = {}
    if not valido:
        erros.update({campo: [str(e) for e in lista] for campo, lista in form.errors.items()})

    if formularios:
        # Erro mostrado com o nome da pergunta (ex.: "Cor favorita"), não
        # a chave interna (ex.: "cor_favorita") — mais legível na revisão.
        # O rótulo é montado uma vez, combinando as variáveis de todos os
        # formulários do perfil (chave já é única globalmente).
        rotulo_por_chave = {fv.variavel.chave: fv.variavel.nome for fv in variaveis_dos_formularios(formularios)}
        for formulario in formularios:
            form_dinamico, _linhas = construir_form_resposta(formulario, data=dados)
            if not form_dinamico.is_valid():
                valido = False
                erros.update(
                    {
                        rotulo_por_chave.get(campo, campo): [str(e) for e in lista]
                        for campo, lista in form_dinamico.errors.items()
                    }
                )

    return {
        "dados": dados,
        "valido": valido,
        "erros": erros or None,
        "consentimento": False,
        "existente_pk": existente.pk if existente else None,
        "existente_codigo": existente.codigo if existente else "",
        "legado": False,
    }


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_projeto(request):
    request.session.pop(WIZ_SESSION_KEY, None)
    if request.method == "POST":
        form = EscolherProjetoWizardForm(request.POST)
        if form.is_valid():
            perfil = form.cleaned_data["perfil"]
            recrutador = form.cleaned_data["recrutador"]
            request.session[WIZ_SESSION_KEY] = {
                "perfil_id": perfil.id if perfil else None,
                "recrutador_id": recrutador.id if recrutador else None,
                "legado": form.cleaned_data["legado"],
                "modo": None,
                "linhas": [],
            }
            return redirect("pessoas:wizard_modo")
    else:
        form = EscolherProjetoWizardForm()
    return render(
        request,
        "pessoas/wizard_projeto.html",
        {"form": form, "steps": WIZ_STEPS, "step_atual": 0},
    )


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_modo(request):
    estado = request.session.get(WIZ_SESSION_KEY)
    if estado is None:
        messages.info(request, "Comece escolhendo o destino dos novos participantes.")
        return redirect("pessoas:wizard_projeto")
    if request.method == "POST":
        modo = request.POST.get("modo")
        if modo not in ("CSV", "MANUAL"):
            messages.error(request, "Escolha uma forma de cadastro para continuar.")
        else:
            estado["modo"] = modo
            request.session[WIZ_SESSION_KEY] = estado
            if modo == "CSV":
                return redirect("pessoas:wizard_dados_csv")
            return redirect("pessoas:wizard_dados_manual")
    return render(
        request,
        "pessoas/wizard_modo.html",
        {"steps": WIZ_STEPS, "step_atual": 1},
    )


def _perfil_e_formularios_do_wizard(estado):
    """Perfil escolhido no passo 1 do wizard (`None` se for "sem perfil") e
    a lista ordenada dos formulários desse perfil (pode ser vazia) — usado
    pra planilha do CSV, pro modelo de exemplo e pra gravar as respostas
    dos formulários na revisão final."""
    if not estado or not estado.get("perfil_id"):
        return None, []
    perfil = (
        Perfil.objects.select_related("projeto")
        .prefetch_related("perfil_formularios__formulario")
        .filter(pk=estado["perfil_id"])
        .first()
    )
    formularios = perfil.formularios_ordenados if perfil else []
    return perfil, formularios


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_dados_csv(request):
    estado = request.session.get(WIZ_SESSION_KEY)
    if estado is None or estado.get("modo") != "CSV":
        return redirect("pessoas:wizard_projeto")
    _perfil, formularios = _perfil_e_formularios_do_wizard(estado)

    if request.method == "POST":
        form = UploadCSVForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                linhas_planilha = ler_planilha(form.cleaned_data["arquivo"], formularios)
            except Exception:
                linhas_planilha = None
                messages.error(
                    request, "Não conseguimos ler esse arquivo. Confira se é um .xlsx ou .csv válido."
                )
            if linhas_planilha is not None and not linhas_planilha:
                messages.error(request, "Não encontramos nenhuma linha de dados nesse arquivo.")
            elif linhas_planilha:
                legado = estado.get("legado", False)
                estado["linhas"] = [
                    _validar_linha_csv(dados, formularios, legado) for dados in linhas_planilha
                ]
                request.session[WIZ_SESSION_KEY] = estado
                return redirect("pessoas:wizard_revisao")
    else:
        form = UploadCSVForm()
    return render(
        request,
        "pessoas/wizard_dados_csv.html",
        {
            "form": form,
            "formularios": formularios,
            "legado": estado.get("legado", False),
            "steps": WIZ_STEPS,
            "step_atual": 2,
        },
    )


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_modelo_csv(request):
    estado = request.session.get(WIZ_SESSION_KEY)
    _perfil, formularios = _perfil_e_formularios_do_wizard(estado)

    cabecalho = list(CABECALHO_MODELO)
    linha_exemplo = list(LINHA_EXEMPLO)
    for fv in variaveis_dos_formularios(formularios):
        rotulo = fv.variavel.nome + (" *" if fv.variavel.obrigatoria else "")
        cabecalho.append(rotulo)
        linha_exemplo.append(_exemplo_variavel(fv.variavel))

    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"

    ws.append(cabecalho)
    for col_idx in range(1, len(cabecalho) + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="C4143F")

    ws.append(linha_exemplo)
    # CPF, CEP e telefone ficam como texto — senão o Excel trata como número
    # e come o zero à esquerda (ex.: CEP "01000-000" viraria "1000-000").
    for coluna_texto in ("CPF", "CEP", "Telefone", "Data de nascimento"):
        col_idx = CABECALHO_MODELO.index(coluna_texto) + 1
        ws.cell(row=2, column=col_idx).number_format = "@"

    for col_idx, titulo in enumerate(cabecalho, start=1):
        maior = max(len(titulo), len(str(linha_exemplo[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = maior + 4
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="modelo_importacao_participantes.xlsx"'
    return response


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_dados_manual(request):
    estado = request.session.get(WIZ_SESSION_KEY)
    if estado is None or estado.get("modo") != "MANUAL":
        return redirect("pessoas:wizard_projeto")

    try:
        quantidade = max(1, min(20, int(request.GET.get("linhas", 5))))
    except ValueError:
        quantidade = 5

    if request.method == "POST":
        FormSet = participante_wizard_formset()
        formset = FormSet(request.POST)
        if formset.is_valid():
            linhas = []
            for f in formset:
                if not f.cleaned_data or not f.cleaned_data.get("nome"):
                    continue
                dados = dict(f.cleaned_data)
                if dados.get("data_nascimento"):
                    dados["data_nascimento"] = dados["data_nascimento"].isoformat()
                # `profissao` é um ModelChoiceField — `cleaned_data` traz a
                # instância de `Profissao`, que não é serializável em JSON
                # (a sessão é salva como JSON). Guarda só o PK, do mesmo
                # jeito que `ler_planilha()` já entrega pro CSV/XLSX — o
                # `ParticipanteWizardForm(data=...)` da revisão aceita o PK
                # normalmente ao religar o form.
                dados["profissao"] = dados["profissao"].pk if dados.get("profissao") else ""
                # CPF duplicado já é barrado pelo próprio formset (campo
                # único no model) antes de chegar aqui — e-mail/telefone não
                # têm essa trava, então checa os três pra já sinalizar na
                # revisão se a linha vai atualizar alguém existente.
                existente = encontrar_participante_existente(
                    dados.get("cpf", ""), dados.get("email", ""), dados.get("telefone", "")
                )
                linhas.append(
                    {
                        "dados": dados,
                        "valido": True,
                        "erros": None,
                        "consentimento": False,
                        "existente_pk": existente.pk if existente else None,
                        "existente_codigo": existente.codigo if existente else "",
                    }
                )
            if not linhas:
                messages.error(request, "Preencha ao menos um participante antes de continuar.")
            else:
                estado["linhas"] = linhas
                request.session[WIZ_SESSION_KEY] = estado
                return redirect("pessoas:wizard_revisao")
    else:
        FormSet = participante_wizard_formset(extra=quantidade)
        formset = FormSet()

    return render(
        request,
        "pessoas/wizard_dados_manual.html",
        {"formset": formset, "quantidade": quantidade, "steps": WIZ_STEPS, "step_atual": 2},
    )


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_revisao(request):
    estado = request.session.get(WIZ_SESSION_KEY)
    if not estado or not estado.get("linhas"):
        messages.info(request, "Nenhum participante para revisar ainda.")
        return redirect("pessoas:wizard_projeto")

    linhas = estado["linhas"]
    perfil, formularios = _perfil_e_formularios_do_wizard(estado)
    # As respostas dos formulários do perfil só existem pra quem veio da
    # planilha (`ler_planilha()` já sabe extrair as colunas certas) — o
    # formset manual não tem campo nenhum pra essas perguntas, então não dá
    # pra exigir/gravar resposta de formulário nesse caminho.
    formularios_da_planilha = formularios if estado.get("modo") == "CSV" else []
    rotulo_por_chave = {
        fv.variavel.chave: fv.variavel.nome for fv in variaveis_dos_formularios(formularios_da_planilha)
    }
    campos_editaveis = list(ParticipanteWizardForm().fields.keys())

    if request.method == "POST":
        versao_lgpd = _versao_lgpd_vigente()
        legado = estado.get("legado", False)
        recrutador_escolhido = None
        if estado.get("recrutador_id"):
            recrutador_escolhido = Usuario.objects.filter(pk=estado["recrutador_id"]).first()
        criados = 0
        atualizados = 0
        restantes = []

        for indice, linha in enumerate(linhas):
            dados = dict(linha["dados"])
            # Aplica as correções feitas na tela de revisão (só as linhas
            # com erro ganham o mini-formulário editável, então uma linha já
            # válida não manda nenhum `dados_<indice>_*` e mantém o que já
            # tinha).
            for campo in campos_editaveis:
                chave_post = f"dados_{indice}-{campo}"
                if chave_post in request.POST:
                    dados[campo] = request.POST.get(chave_post, "")

            aceite = request.POST.get(f"consentimento_{indice}") == "on"
            dados["situacao"] = dados.get("situacao") or Participante.Situacao.PENDENTE

            if legado:
                dados_prontos, campos_incompletos, erro_legado = preparar_linha_legado(dados)
                if erro_legado:
                    restantes.append(
                        {
                            "dados": dados,
                            "valido": False,
                            "erros": {"Linha": [erro_legado]},
                            "consentimento": aceite,
                            "existente_pk": None,
                            "existente_codigo": "",
                            "legado": True,
                        }
                    )
                    continue
                dados = dados_prontos
                existente = encontrar_participante_existente(
                    dados.get("cpf") or "", dados.get("email", ""), dados.get("telefone") or ""
                )
                if not aceite:
                    restantes.append(
                        {
                            "dados": dados,
                            "valido": True,
                            "erros": {"Consentimento": ["Marque o consentimento LGPD pra importar esta linha."]},
                            "consentimento": False,
                            "existente_pk": existente.pk if existente else None,
                            "existente_codigo": existente.codigo if existente else "",
                            "legado": True,
                        }
                    )
                    continue

                valores_originais = capturar_valores_atuais(existente) if existente else None
                participante = existente or Participante()
                aplicar_dados_legado(participante, dados, campos_incompletos, existente, valores_originais)
                forms_dinamicos = []
            else:
                existente = encontrar_participante_existente(
                    dados.get("cpf", ""), dados.get("email", ""), dados.get("telefone", "")
                )
                valores_originais = capturar_valores_atuais(existente) if existente else None
                form = ParticipanteWizardForm(data=dados, instance=existente)
                valido = form.is_valid()

                # Um form dinâmico por formulário do perfil — cada um vira
                # uma `RespostaFormulario` separada depois de salvar.
                forms_dinamicos = []
                for formulario in formularios_da_planilha:
                    form_dinamico, _campos_dinamicos = construir_form_resposta(formulario, data=dados)
                    forms_dinamicos.append((formulario, form_dinamico))
                    valido = valido and form_dinamico.is_valid()

                if not (valido and aceite):
                    erros = {}
                    if not form.is_valid():
                        erros.update({c: [str(e) for e in lst] for c, lst in form.errors.items()})
                    for _formulario, form_dinamico in forms_dinamicos:
                        if not form_dinamico.is_valid():
                            erros.update(
                                {
                                    rotulo_por_chave.get(c, c): [str(e) for e in lst]
                                    for c, lst in form_dinamico.errors.items()
                                }
                            )
                    if valido and not aceite:
                        erros["Consentimento"] = ["Marque o consentimento LGPD pra importar esta linha."]
                    restantes.append(
                        {
                            "dados": dados,
                            "valido": False,
                            "erros": erros or None,
                            "consentimento": aceite,
                            "existente_pk": existente.pk if existente else None,
                            "existente_codigo": existente.codigo if existente else "",
                            "legado": False,
                        }
                    )
                    continue

                participante = form.save(commit=False)
                if existente:
                    # Atualização: uma submissão com campo vazio não apaga o
                    # que já tinha sido salvo.
                    restaurar_campos_vazios(participante, form.cleaned_data, valores_originais)

            if existente:
                novo_consentimento = not existente.consentimento_versao_id
            else:
                participante.criado_por = request.user
                # Recrutador escolhido no passo 1 do wizard — se ninguém foi
                # escolhido, quem está enviando o lote assume esse papel.
                # Só se aplica a cadastro novo (atualizar não rouba o
                # crédito de quem indicou a pessoa da primeira vez).
                participante.origem_recrutador = recrutador_escolhido or request.user
                novo_consentimento = True
            if novo_consentimento:
                participante.consentimento_lgpd = True
                participante.consentimento_versao = versao_lgpd
            participante.save()
            if novo_consentimento and versao_lgpd:
                registrar_aceite(
                    participante, versao_lgpd,
                    origem=AceiteTermo.Origem.STAFF, request=request, registrado_por=request.user,
                )
            if existente:
                atualizados += 1
            else:
                criados += 1

            if perfil is not None:
                participacao, _criada = Participacao.objects.get_or_create(
                    participante=participante,
                    perfil=perfil,
                    defaults={"etapa": Participacao.Etapa.ANALISE_PERFIL, "responsavel": request.user},
                )
                for formulario, form_dinamico in forms_dinamicos:
                    RespostaFormulario.objects.update_or_create(
                        participacao=participacao,
                        formulario=formulario,
                        defaults={"respostas_variaveis": form_dinamico.cleaned_data},
                    )

        if criados:
            messages.success(request, f"{criados} participante(s) importado(s) com sucesso.")
        if atualizados:
            messages.success(request, f"{atualizados} participante(s) já cadastrado(s) foram atualizados.")

        if restantes:
            estado["linhas"] = restantes
            request.session[WIZ_SESSION_KEY] = estado
            messages.info(
                request,
                f"{len(restantes)} linha(s) ainda precisam de ajuste — corrija os campos "
                'destacados e clique em "Concluir importação" de novo.',
            )
        else:
            del request.session[WIZ_SESSION_KEY]
            if not criados and not atualizados:
                messages.info(request, "Nenhuma linha foi processada.")
            return redirect("pessoas:lista")

    linhas_exibicao = []
    for indice, linha in enumerate(linhas):
        linha_exibicao = {"indice": indice, **linha}
        if not linha["valido"]:
            # Só linha com erro ganha o mini-formulário editável — reaproveita
            # o mesmo ParticipanteWizardForm do cadastro manual, já pré-cheio
            # com o que a planilha trouxe, pra corrigir sem perder o resto.
            linha_exibicao["form"] = ParticipanteWizardForm(initial=linha["dados"], prefix=f"dados_{indice}")
        linhas_exibicao.append(linha_exibicao)
    return render(
        request,
        "pessoas/wizard_revisao.html",
        {
            "linhas": linhas_exibicao,
            "perfil": perfil,
            "campos_editaveis": campos_editaveis,
            "legado": estado.get("legado", False),
            "steps": WIZ_STEPS,
            "step_atual": 3,
        },
    )


@login_required
@requer_permissao("participantes.gerenciar")
def wizard_cancelar(request):
    request.session.pop(WIZ_SESSION_KEY, None)
    messages.info(request, "Importação cancelada.")
    return redirect("pessoas:lista")


# =========================================================================
# Triagem (aprovar/descartar participantes pendentes)
# =========================================================================


@login_required
@requer_permissao("participantes.gerenciar")
@require_POST
def aprovar(request, pk):
    participante = get_object_or_404(Participante, pk=pk)
    participante.situacao = Participante.Situacao.APROVADO
    participante.save(update_fields=["situacao"])
    registrar(request.user, participante.codigo, RegistroAcesso.Acao.ALTERACAO, "Participante aprovado na triagem")
    messages.success(request, f"{participante.nome} aprovado(a).")
    return redirect("pessoas:detalhe", pk=pk)


@login_required
@requer_permissao("participantes.gerenciar")
@require_POST
def descartar(request, pk):
    participante = get_object_or_404(Participante, pk=pk)
    participante.situacao = Participante.Situacao.DESCARTADO
    participante.save(update_fields=["situacao"])
    # Limpa participações que ainda não passaram de revisão inicial — tanto
    # quem acabou de se cadastrar pelo link (Preenchimento de Dados) quanto
    # quem já foi pra Análise de Perfil sem ainda ter avançado mais no
    # funil. Quem já passou dessas duas etapas fica intocado (descartar o
    # participante não deveria apagar um histórico de funil já avançado).
    Participacao.objects.filter(
        participante=participante,
        etapa__in=[Participacao.Etapa.PREENCHIMENTO_DADOS, Participacao.Etapa.ANALISE_PERFIL],
    ).delete()
    registrar(request.user, participante.codigo, RegistroAcesso.Acao.ALTERACAO, "Participante descartado na triagem")
    messages.info(request, f"{participante.nome} descartado(a).")
    return redirect("pessoas:detalhe", pk=pk)


# =========================================================================
# Página pública de cadastro (sem login) — acessada por link de recrutador
# =========================================================================


def _categorias_disponiveis_para_escolha(perfil):
    """Categorias distintas entre os formulários ativos do perfil, em ordem
    alfabética — a lista que a pessoa vê na tela de escolha. Formulário sem
    categoria não entra aqui (esses sempre aparecem pra todo mundo, ver
    `_form_dinamico_do_perfil`)."""
    vistas = {}
    for formulario in perfil.formularios_ordenados:
        if not formulario.ativo or formulario.categoria_id is None:
            continue
        vistas.setdefault(formulario.categoria_id, formulario.categoria)
    return sorted(vistas.values(), key=lambda categoria: categoria.nome)


def _form_dinamico_do_perfil(perfil, data=None, categorias_ids=None):
    """Devolve uma lista com um item por formulário ativo do perfil (na
    ordem escolhida na tela do perfil), cada um já montado por
    `formularios/respostas.py::construir_form_resposta` — pronta pra
    iterar num único `{% for %}` no template.

    `categorias_ids`, quando não é `None`, restringe aos formulários cuja
    categoria está nesse conjunto (formulário sem categoria sempre entra,
    categorizar é opcional e não deveria esconder pergunta nenhuma). É
    `None` sempre que o perfil não exige escolha de categorias (perfil de
    Respostas, ou perfil de Captação com poucas categorias — ver
    `Perfil.qtd_categorias_escolha`)."""
    itens = []
    for formulario in perfil.formularios_ordenados:
        if not formulario.ativo:
            continue
        if categorias_ids is not None and formulario.categoria_id is not None:
            if str(formulario.categoria_id) not in categorias_ids:
                continue
        form_dinamico, linhas = construir_form_resposta(formulario, data=data)
        itens.append({"formulario": formulario, "form": form_dinamico, "linhas": linhas})
    return itens


def cadastro_publico(request, token):
    try:
        payload = ler_token_captacao(token)
    except ValueError as exc:
        return render(request, "publico/link_invalido.html", {"mensagem": str(exc)}, status=410)

    # O link em si não expira mais por tempo — só funciona enquanto o projeto
    # está "Recrutando". Sai desse status (em campo, concluído, o que for) e o
    # link para de aceitar novos cadastros sozinho, sem precisar gerar outro.
    perfil = (
        Perfil.objects.select_related("projeto")
        .prefetch_related("perfil_formularios__formulario")
        .filter(pk=payload["perfil_id"], projeto__status=Projeto.Status.RECRUTANDO)
        .first()
    )
    if perfil is None:
        return render(
            request,
            "publico/link_invalido.html",
            {"mensagem": "Este projeto não está mais recebendo cadastros."},
            status=410,
        )
    projeto = perfil.projeto

    recrutador = Usuario.objects.filter(pk=payload["recrutador_id"]).first()
    versao_lgpd = _versao_lgpd_vigente()

    # Perfil de Captação com mais categorias de formulário associadas do que
    # `qtd_categorias_escolha`: a pessoa escolhe essa quantidade antes de ver
    # qualquer pergunta, e só os formulários dessas categorias (mais os sem
    # categoria) abrem pra responder — ver
    # `_categorias_disponiveis_para_escolha`/`_form_dinamico_do_perfil`.
    # Quantidade, pergunta e texto de boas-vindas são configurados por
    # perfil (`projetos/forms.py::PerfilForm`), não fixos no código.
    categorias_disponiveis = _categorias_disponiveis_para_escolha(perfil)
    qtd_categorias = perfil.qtd_categorias_escolha
    exige_escolha_categorias = perfil.tipo == Perfil.Tipo.CAPTACAO and len(categorias_disponiveis) > qtd_categorias
    categorias_ids_validos = {str(categoria.id) for categoria in categorias_disponiveis}

    def _categorias_selecionadas(fonte):
        ids = [cid for cid in fonte.getlist("categorias") if cid in categorias_ids_validos]
        # dedup preservando ordem, caso o mesmo id venha repetido
        return list(dict.fromkeys(ids))

    def _contexto_escolha_categorias(erro=None):
        contexto = {
            "projeto": projeto, "perfil": perfil, "categorias": categorias_disponiveis,
            "num_categorias": qtd_categorias,
            "texto_pergunta": perfil.texto_escolha_categorias_efetivo,
            "texto_boas_vindas": perfil.texto_boas_vindas_categorias,
        }
        if erro:
            contexto["erro"] = erro
        return contexto

    if request.method == "POST":
        categorias_selecionadas = _categorias_selecionadas(request.POST) if exige_escolha_categorias else []
        if exige_escolha_categorias and len(categorias_selecionadas) != qtd_categorias:
            # POST forjado/adulterado sem a quantidade certa de categorias —
            # volta pra tela de escolha em vez de tentar validar um
            # formulário incompleto.
            return render(
                request,
                "publico/escolha_categorias.html",
                _contexto_escolha_categorias(erro=f"Selecione exatamente {qtd_categorias} categorias."),
            )
        existente = encontrar_participante_existente(
            request.POST.get("cpf", ""), request.POST.get("email", ""), request.POST.get("telefone", "")
        )
        valores_originais = capturar_valores_atuais(existente) if existente else None
        form = CadastroPublicoForm(request.POST, instance=existente)
        forms_dinamicos = _form_dinamico_do_perfil(
            perfil, data=request.POST, categorias_ids=categorias_selecionadas if exige_escolha_categorias else None
        )
        valido = form.is_valid() and all(fd["form"].is_valid() for fd in forms_dinamicos)
        if valido:
            with transaction.atomic():
                participante = form.save(commit=False)
                if existente:
                    restaurar_campos_vazios(participante, form.cleaned_data, valores_originais)
                else:
                    # Só um cadastro público novo entra automaticamente como
                    # "Pendente" com esse recrutador como origem — reenviar o
                    # mesmo formulário pra atualizar dados não deve desfazer
                    # uma triagem já feita nem trocar quem indicou a pessoa.
                    participante.situacao = Participante.Situacao.PENDENTE
                    participante.origem_recrutador = recrutador
                if versao_lgpd and not participante.consentimento_versao_id:
                    participante.consentimento_versao = versao_lgpd
                participante.save()
                if versao_lgpd and not (existente and existente.consentimento_versao_id):
                    registrar_aceite(
                        participante, versao_lgpd,
                        origem=AceiteTermo.Origem.PUBLICO, request=request, registrado_por=recrutador,
                    )
                # Etapa inicial é "Preenchimento de Dados" — é literalmente
                # o que acabou de acontecer aqui (a pessoa preencheu os
                # dados pelo link público). A equipe analisa o perfil depois,
                # como próximo passo do funil.
                participacao, _criada = Participacao.objects.get_or_create(
                    participante=participante,
                    perfil=perfil,
                    defaults={"etapa": Participacao.Etapa.PREENCHIMENTO_DADOS, "responsavel": recrutador},
                )
                for fd in forms_dinamicos:
                    if not fd["linhas"]:
                        continue
                    RespostaFormulario.objects.update_or_create(
                        participacao=participacao,
                        formulario=fd["formulario"],
                        defaults={"respostas_variaveis": fd["form"].cleaned_data},
                    )
            return render(
                request, "publico/cadastro_ok.html", {"participante": participante, "projeto": projeto, "perfil": perfil}
            )
    else:
        categorias_selecionadas = _categorias_selecionadas(request.GET) if exige_escolha_categorias else []
        if exige_escolha_categorias and len(categorias_selecionadas) != qtd_categorias:
            return render(request, "publico/escolha_categorias.html", _contexto_escolha_categorias())
        form = CadastroPublicoForm()
        forms_dinamicos = _form_dinamico_do_perfil(
            perfil, categorias_ids=categorias_selecionadas if exige_escolha_categorias else None
        )

    return render(
        request,
        "publico/cadastro.html",
        {
            "form": form,
            "projeto": projeto,
            "perfil": perfil,
            "versao_lgpd": versao_lgpd,
            "forms_dinamicos": forms_dinamicos,
            "categorias_selecionadas": categorias_selecionadas if exige_escolha_categorias else None,
        },
    )


def renovar_termo(request, token):
    """Tela pública (sem login) de renovação de um termo/contrato — link
    gerado em `pessoas:detalhe` (ver `_link_renovacao_termo`) pra reenviar
    pra uma pessoa que já aceitou uma versão antiga do documento. Identidade
    confirmada pelo CPF (tem que bater com o cadastro); e-mail é só
    informativo, de propósito não trava o aceite (a mesma pessoa pode ter
    mais de um e-mail com o tempo)."""
    try:
        payload = ler_token_renovacao_termo(token)
    except ValueError as exc:
        return render(request, "publico/link_invalido.html", {"mensagem": str(exc)}, status=410)

    participante = Participante.objects.filter(pk=payload["participante_id"]).first()
    termo = Termo.objects.filter(pk=payload["termo_id"]).first()
    if participante is None or termo is None:
        return render(
            request,
            "publico/link_invalido.html",
            {"mensagem": "Este link de renovação não é mais válido."},
            status=410,
        )

    versao_vigente = termo.versao_vigente
    if versao_vigente is None:
        return render(
            request,
            "publico/link_invalido.html",
            {"mensagem": "Este documento não tem uma versão vigente no momento."},
            status=410,
        )

    ja_esta_em_dia = AceiteTermo.objects.filter(participante=participante, versao=versao_vigente).exists()
    if ja_esta_em_dia:
        return render(
            request,
            "publico/termo_renovado_ok.html",
            {"participante": participante, "termo": termo, "versao": versao_vigente, "ja_estava_em_dia": True},
        )

    if request.method == "POST":
        form = RenovarTermoForm(request.POST)
        if form.is_valid():
            cpf_confere = normalizar_cpf(form.cleaned_data["cpf"]) == normalizar_cpf(participante.cpf or "")
            if not cpf_confere:
                form.add_error("cpf", "Esse CPF não confere com o cadastro — confira e tente de novo.")
            else:
                registrar_aceite(
                    participante, versao_vigente, origem=AceiteTermo.Origem.PUBLICO, request=request
                )
                # `consentimento_versao` só existe pro termo de Consentimento
                # LGPD especificamente (é o único documento com esse ponteiro
                # dedicado em Participante) — outros tipos de termo/contrato
                # ficam só no histórico de `AceiteTermo`, sem campo próprio.
                termo_lgpd = Termo.objects.filter(tipo=Termo.Tipo.CONSENTIMENTO).order_by("id").first()
                if termo_lgpd and termo_lgpd.pk == termo.pk:
                    participante.consentimento_versao = versao_vigente
                    participante.consentimento_lgpd = True
                    participante.save(update_fields=["consentimento_versao", "consentimento_lgpd"])
                return render(
                    request,
                    "publico/termo_renovado_ok.html",
                    {"participante": participante, "termo": termo, "versao": versao_vigente},
                )
    else:
        form = RenovarTermoForm()

    return render(
        request,
        "publico/renovar_termo.html",
        {"participante": participante, "termo": termo, "versao": versao_vigente, "form": form},
    )
