"""Exportação da lista de Pessoas em PDF e XLSX.

Diferente da exportação de Participações, aqui NUNCA entram dados sensíveis
(CPF, telefone, e-mail, forma de pagamento/chave PIX) — nem mascarados —,
independente da permissão de quem está baixando: é um relatório de perfil
agregado, não um jeito de tirar a base de contato de dentro do sistema. Por
isso também é limitada a 50 pessoas por download (ver `LIMITE_LINHAS`) e o
próprio acesso à view já passa por um limite de 1 exportação a cada 12h por
usuário (`auditoria.verificar_limite_download`)."""

import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from core.relatorios import caminho_logo

LIMITE_LINHAS = 50

VIOLETA_DEEP = colors.HexColor("#C4143F")
VIOLETA_SOFT = colors.HexColor("#FDE8EE")
TINTA = colors.HexColor("#251A1E")
MUDO = colors.HexColor("#9A8790")
LINHA = colors.HexColor("#F1E4E8")


def _linhas(participantes):
    linhas = []
    for p in participantes:
        linhas.append(
            {
                "codigo": p.codigo,
                "nome": p.nome,
                "cidade": f"{p.cidade}/{p.uf}",
                "idade": f"{p.idade} anos",
                "escolaridade": p.get_escolaridade_display() or "—",
                "classe_social": p.get_faixa_renda_display() or "—",
                "situacao": p.get_situacao_display(),
                "ultima_participacao": (
                    p.data_ultima_participacao.strftime("%d/%m/%Y") if p.data_ultima_participacao else "—"
                ),
                "cadastrado_em": timezone.localtime(p.criado_em).strftime("%d/%m/%Y"),
                "consentimento": "Aceito" if p.consentimento_lgpd else "Não aceito",
                "cadastro_incompleto": "Incompleto" if p.cadastro_incompleto else "Completo",
            }
        )
    return linhas


def _resumo_filtros(filtros):
    partes = []
    if filtros.get("q"):
        partes.append(f"busca “{filtros['q']}”")
    if filtros.get("situacao"):
        partes.append(f"situação: {filtros['situacao_label']}")
    if filtros.get("faixa_renda"):
        partes.append(f"classe social: {filtros['faixa_renda_label']}")
    if filtros.get("uf"):
        partes.append(f"UF: {filtros['uf']}")
    return "; ".join(partes) if partes else "nenhum filtro aplicado — todas as pessoas"


def gerar_xlsx(participantes, filtros):
    linhas = _linhas(participantes[:LIMITE_LINHAS])
    total = participantes.count()

    colunas = [
        ("codigo", "Código"),
        ("nome", "Nome"),
        ("cidade", "Cidade/UF"),
        ("idade", "Idade"),
        ("escolaridade", "Escolaridade"),
        ("classe_social", "Classe social"),
        ("situacao", "Situação"),
        ("ultima_participacao", "Última participação"),
        ("cadastrado_em", "Cadastrado em"),
        ("consentimento", "Consentimento LGPD"),
        ("cadastro_incompleto", "Cadastro"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pessoas"

    ws.append(["Qualy Vortice — Relatório de Pessoas"])
    ws["A1"].font = Font(bold=True, size=14, color="C4143F")
    ws.append([f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"])
    ws["A2"].font = Font(size=9, color="9A8790")
    ws.append([f"Mostrando {len(linhas)} de {total} pessoa(s) — limite de {LIMITE_LINHAS} por download"])
    ws["A3"].font = Font(size=9, color="9A8790")
    ws.append(["Este relatório não inclui CPF, telefone, e-mail nem dados de pagamento."])
    ws["A4"].font = Font(size=9, italic=True, color="9A8790")
    ws.append([])

    cabecalho_linha = 6
    ws.append([rotulo for _, rotulo in colunas])
    for col_idx in range(1, len(colunas) + 1):
        celula = ws.cell(row=cabecalho_linha, column=col_idx)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="C4143F")
        celula.alignment = Alignment(vertical="center", wrap_text=True)

    for linha in linhas:
        ws.append([linha[chave] for chave, _ in colunas])

    for col_idx, (chave, rotulo) in enumerate(colunas, start=1):
        maior = max([len(rotulo)] + [len(str(linha[chave])) for linha in linhas] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(maior + 2, 10), 40)

    ws.freeze_panes = f"A{cabecalho_linha + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    nome_arquivo = f"pessoas_{timezone.localtime().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def gerar_pdf(participantes, filtros):
    linhas = _linhas(participantes[:LIMITE_LINHAS])
    total = participantes.count()

    estilos = getSampleStyleSheet()
    estilo_celula = ParagraphStyle("celula", parent=estilos["Normal"], fontSize=8, leading=9.5, textColor=TINTA)
    estilo_titulo = ParagraphStyle(
        "titulo", parent=estilos["Heading1"], fontSize=16, textColor=VIOLETA_DEEP, spaceAfter=2
    )
    estilo_sub = ParagraphStyle("sub", parent=estilos["Normal"], fontSize=8.5, textColor=MUDO)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.4 * cm,
        title="Relatório de Pessoas — Qualy Vortice",
    )

    elementos = []

    logo = caminho_logo()
    if logo:
        cabecalho = Table(
            [[Image(logo, width=1.6 * cm, height=1.6 * cm), Paragraph("Qualy Vortice", estilo_titulo)]],
            colWidths=[1.9 * cm, None],
        )
        cabecalho.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elementos.append(cabecalho)
    else:
        elementos.append(Paragraph("Qualy Vortice", estilo_titulo))

    elementos.append(Paragraph("Relatório de Pessoas", estilos["Heading2"]))
    elementos.append(Paragraph(f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}", estilo_sub))
    elementos.append(Paragraph(f"Filtros: {_resumo_filtros(filtros)}", estilo_sub))
    elementos.append(
        Paragraph(f"Mostrando {len(linhas)} de {total} pessoa(s) — limite de {LIMITE_LINHAS} por download", estilo_sub)
    )
    elementos.append(
        Paragraph("Este relatório não inclui CPF, telefone, e-mail nem dados de pagamento.", estilo_sub)
    )
    elementos.append(Spacer(1, 10))

    cabecalhos = [
        "Código", "Nome", "Cidade/UF", "Idade", "Escolaridade", "Classe social",
        "Situação", "Última participação", "Cadastrado em", "Consentimento LGPD", "Cadastro",
    ]
    dados = [
        [
            Paragraph(c, ParagraphStyle("cab", parent=estilo_celula, textColor=colors.white, fontName="Helvetica-Bold"))
            for c in cabecalhos
        ]
    ]
    for linha in linhas:
        dados.append(
            [
                Paragraph(linha["codigo"], estilo_celula),
                Paragraph(linha["nome"], estilo_celula),
                Paragraph(linha["cidade"], estilo_celula),
                Paragraph(linha["idade"], estilo_celula),
                Paragraph(linha["escolaridade"], estilo_celula),
                Paragraph(linha["classe_social"], estilo_celula),
                Paragraph(linha["situacao"], estilo_celula),
                Paragraph(linha["ultima_participacao"], estilo_celula),
                Paragraph(linha["cadastrado_em"], estilo_celula),
                Paragraph(linha["consentimento"], estilo_celula),
                Paragraph(linha["cadastro_incompleto"], estilo_celula),
            ]
        )

    largura_util = landscape(A4)[0] - 2.4 * cm
    proporcoes = [0.08, 0.16, 0.11, 0.06, 0.09, 0.09, 0.08, 0.09, 0.07, 0.09, 0.08]
    larguras = [largura_util * p for p in proporcoes]

    tabela = Table(dados, colWidths=larguras, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), VIOLETA_DEEP),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, VIOLETA_SOFT]),
                ("GRID", (0, 0), (-1, -1), 0.4, LINHA),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elementos.append(tabela)

    def _rodape(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(MUDO)
        canvas.drawString(1.2 * cm, 0.8 * cm, "Qualy Vortice — documento gerado pelo sistema")
        canvas.drawRightString(landscape(A4)[0] - 1.2 * cm, 0.8 * cm, f"Página {_doc.page}")
        canvas.restoreState()

    doc.build(elementos, onFirstPage=_rodape, onLaterPages=_rodape)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    nome_arquivo = f"pessoas_{timezone.localtime().strftime('%Y%m%d_%H%M')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response
