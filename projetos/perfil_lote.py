"""Associação em lote de participantes já cadastrados no Banco de Pessoas a
um Perfil, via planilha com uma única coluna de CPF — diferente do wizard de
`pessoas` (que importa/cadastra participantes novos), aqui a pessoa já
existe: só falta ligar ela a um perfil."""

import csv
import io

from django.db.models import Value
from django.db.models.functions import Replace
from openpyxl import load_workbook

from participacoes.models import Participacao
from pessoas.models import Participante
from pessoas.validators import normalizar_cpf


def _texto_celula(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _ler_csv(arquivo):
    bruto = arquivo.read()
    for codificacao in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = bruto.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue
    else:
        texto = bruto.decode("utf-8", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(texto[:2000], delimiters=",;")
    except csv.Error:
        dialect = csv.excel

    leitor = csv.DictReader(io.StringIO(texto), dialect=dialect)
    valores = []
    for linha in leitor:
        for chave, valor in linha.items():
            if (chave or "").strip().lower() == "cpf":
                valores.append(valor or "")
                break
    return valores


def _ler_xlsx(arquivo):
    pasta = load_workbook(arquivo, data_only=True, read_only=True)
    aba = pasta.active
    linhas = aba.iter_rows(values_only=True)
    try:
        cabecalho = [(str(c or "")).strip().lower() for c in next(linhas)]
    except StopIteration:
        return []
    if "cpf" not in cabecalho:
        return []
    indice = cabecalho.index("cpf")
    return [_texto_celula(linha[indice]) for linha in linhas if indice < len(linha)]


def ler_cpfs(arquivo):
    """Lê uma planilha (.xlsx ou .csv) com uma coluna "CPF" e devolve a
    lista de CPFs normalizados (só dígitos), sem duplicatas e sem vazios."""
    nome = (arquivo.name or "").lower()
    valores = _ler_xlsx(arquivo) if nome.endswith(".xlsx") else _ler_csv(arquivo)
    normalizados = [normalizar_cpf(v) for v in valores]
    vistos = []
    for cpf in normalizados:
        if cpf and cpf not in vistos:
            vistos.append(cpf)
    return vistos


def associar_cpfs(perfil, cpfs, usuario):
    """Associa cada CPF encontrado no Banco de Pessoas ao `perfil` (cria uma
    Participacao se ainda não existir uma pra esse par). Devolve um resumo
    pra tela: quem foi associado agora, quem já estava, e os CPFs que não
    bateram com nenhum Participante cadastrado."""
    participantes_por_cpf = {
        p.cpf_normalizado: p
        for p in Participante.objects.annotate(
            cpf_normalizado=Replace(Replace("cpf", Value("."), Value("")), Value("-"), Value(""))
        ).filter(cpf_normalizado__in=cpfs)
    }

    associados, ja_associados, nao_encontrados = [], [], []
    for cpf in cpfs:
        participante = participantes_por_cpf.get(cpf)
        if not participante:
            nao_encontrados.append(cpf)
            continue
        _participacao, criada = Participacao.objects.get_or_create(
            participante=participante,
            perfil=perfil,
            defaults={"etapa": Participacao.Etapa.ANALISE_PERFIL, "responsavel": usuario},
        )
        (associados if criada else ja_associados).append(participante)

    return {"associados": associados, "ja_associados": ja_associados, "nao_encontrados": nao_encontrados}
