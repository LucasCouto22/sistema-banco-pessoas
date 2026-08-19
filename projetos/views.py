import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from accounts.permissions import requer_permissao
from pessoas.forms import UploadCSVForm
from pessoas.links import gerar_token_captacao

from .forms import PerfilForm, ProjetoForm, montar_formset_formularios_perfil
from .models import Perfil, PerfilFormulario, Projeto
from .perfil_lote import associar_cpfs, ler_cpfs


@login_required
@requer_permissao("projetos.ver")
def lista(request):
    projetos = Projeto.objects.all()
    return render(request, "projetos/lista.html", {"projetos": projetos})


def _link_captacao(request, perfil):
    """Link público de cadastro do perfil — não precisa mais de tela própria
    pra "gerar": o token não expira por tempo (só perde validade se o
    projeto sair de "Recrutando"), então dá pra montar na hora, direto onde
    o link é mostrado. `recrutador_id` é sempre quem está vendo a página
    nesse momento — cada indicação feita a partir desse link fica atribuída
    a quem copiou, não a quem criou o perfil."""
    token = gerar_token_captacao(perfil.id, request.user.id)
    return request.build_absolute_uri(reverse("pessoas:cadastro_publico", args=[token]))


@login_required
@requer_permissao("projetos.ver")
def detalhe(request, pk):
    projeto = get_object_or_404(Projeto, pk=pk)
    perfis = list(projeto.perfis.prefetch_related("perfil_formularios__formulario"))
    for perfil in perfis:
        perfil.link_cadastro = _link_captacao(request, perfil)
    return render(
        request,
        "projetos/detalhe.html",
        {
            "projeto": projeto,
            "perfis": perfis,
            "pode_editar": request.user.tem_permissao("projetos.gerenciar"),
            "pode_excluir": request.user.tem_permissao("projetos.excluir"),
            "pode_associar": request.user.tem_permissao("participacoes.mover_etapa"),
        },
    )


@login_required
@requer_permissao("projetos.gerenciar")
def novo(request):
    if request.method == "POST":
        form = ProjetoForm(request.POST)
        if form.is_valid():
            projeto = form.save(commit=False)
            projeto.criado_por = request.user
            projeto.save()
            messages.success(request, f"Projeto {projeto.nome} criado.")
            return redirect("projetos:detalhe", pk=projeto.pk)
    else:
        form = ProjetoForm()
    return render(request, "projetos/form.html", {"form": form, "titulo": "Novo projeto"})


@login_required
@requer_permissao("projetos.gerenciar")
def editar(request, pk):
    projeto = get_object_or_404(Projeto, pk=pk)
    if request.method == "POST":
        form = ProjetoForm(request.POST, instance=projeto)
        if form.is_valid():
            form.save()
            messages.success(request, "Projeto atualizado.")
            return redirect("projetos:detalhe", pk=projeto.pk)
    else:
        form = ProjetoForm(instance=projeto)
    return render(request, "projetos/form.html", {"form": form, "titulo": f"Editar {projeto.nome}"})


@login_required
@requer_permissao("projetos.excluir")
def excluir(request, pk):
    projeto = get_object_or_404(Projeto, pk=pk)
    if request.method == "POST":
        nome = projeto.nome
        projeto.delete()
        messages.success(request, f"Projeto {nome} excluído.")
        return redirect("projetos:lista")
    return render(request, "projetos/excluir.html", {"projeto": projeto})


# =========================================================================
# Perfis — de 1 a N por projeto, cada um com seu formulário e seus
# participantes. Geridos a partir da tela de detalhe do projeto (já salvo),
# não embutidos no formulário de criação — mesmo padrão de Formulário e
# Variável, que também são CRUDs à parte.
# =========================================================================


def _salvar_perfil_com_formularios(form, formset, projeto=None):
    """Espelha `formularios/views.py::_salvar_formulario_com_variaveis` —
    salva o Perfil e sincroniza seus `PerfilFormulario` (inclui/atualiza
    ordem dos marcados, remove os desmarcados) numa transação só."""
    if not (form.is_valid() and formset.is_valid()):
        return False

    with transaction.atomic():
        perfil = form.save(commit=False)
        if projeto is not None:
            perfil.projeto = projeto
        perfil.save()

        incluidos = {}
        for subform in formset.forms:
            if not subform.cleaned_data.get("incluir"):
                continue
            incluidos[subform.cleaned_data["formulario_id"]] = subform.cleaned_data.get("ordem") or 0

        PerfilFormulario.objects.filter(perfil=perfil).exclude(
            formulario_id__in=incluidos.keys()
        ).delete()
        for formulario_id, ordem in incluidos.items():
            PerfilFormulario.objects.update_or_create(
                perfil=perfil, formulario_id=formulario_id, defaults={"ordem": ordem}
            )
    return perfil


@login_required
@requer_permissao("projetos.gerenciar")
def perfil_novo(request, projeto_pk):
    projeto = get_object_or_404(Projeto, pk=projeto_pk)
    if request.method == "POST":
        form = PerfilForm(request.POST)
        linhas, formset = montar_formset_formularios_perfil(data=request.POST)
        perfil = _salvar_perfil_com_formularios(form, formset, projeto=projeto)
        if perfil:
            messages.success(request, f'Perfil "{perfil.nome}" criado.')
            return redirect("projetos:detalhe", pk=projeto.pk)
    else:
        form = PerfilForm()
        linhas, formset = montar_formset_formularios_perfil()
    return render(
        request,
        "projetos/perfil_form.html",
        {"form": form, "formset": formset, "linhas": linhas, "projeto": projeto, "titulo": "Novo perfil"},
    )


@login_required
@requer_permissao("projetos.gerenciar")
def perfil_editar(request, pk):
    perfil = get_object_or_404(Perfil.objects.select_related("projeto"), pk=pk)
    if request.method == "POST":
        form = PerfilForm(request.POST, instance=perfil)
        linhas, formset = montar_formset_formularios_perfil(perfil=perfil, data=request.POST)
        if _salvar_perfil_com_formularios(form, formset):
            messages.success(request, "Perfil atualizado.")
            return redirect("projetos:detalhe", pk=perfil.projeto_id)
    else:
        form = PerfilForm(instance=perfil)
        linhas, formset = montar_formset_formularios_perfil(perfil=perfil)
    return render(
        request,
        "projetos/perfil_form.html",
        {
            "form": form,
            "formset": formset,
            "linhas": linhas,
            "projeto": perfil.projeto,
            "titulo": f"Editar {perfil.nome}",
        },
    )


@login_required
@requer_permissao("projetos.gerenciar")
def perfil_excluir(request, pk):
    perfil = get_object_or_404(Perfil.objects.select_related("projeto"), pk=pk)
    if request.method == "POST":
        projeto_pk = perfil.projeto_id
        nome = perfil.nome
        perfil.delete()
        messages.success(request, f'Perfil "{nome}" excluído.')
        return redirect("projetos:detalhe", pk=projeto_pk)
    return render(request, "projetos/perfil_excluir.html", {"perfil": perfil})


@login_required
@requer_permissao("projetos.ver")
def perfil_detalhe(request, pk):
    perfil = get_object_or_404(
        Perfil.objects.select_related("projeto").prefetch_related("perfil_formularios__formulario"), pk=pk
    )
    participacoes = perfil.participacoes.select_related("participante", "responsavel")
    perfil.link_cadastro = _link_captacao(request, perfil)
    return render(
        request,
        "projetos/perfil_detalhe.html",
        {
            "perfil": perfil,
            "participacoes": participacoes,
            "pode_editar": request.user.tem_permissao("projetos.gerenciar"),
            "pode_associar": request.user.tem_permissao("participacoes.mover_etapa"),
        },
    )


@login_required
@requer_permissao("participacoes.mover_etapa")
def perfil_associar_lote(request, pk):
    perfil = get_object_or_404(Perfil.objects.select_related("projeto"), pk=pk)
    resultado = None
    if request.method == "POST":
        form = UploadCSVForm(request.POST, request.FILES)
        if form.is_valid():
            cpfs = ler_cpfs(form.cleaned_data["arquivo"])
            if not cpfs:
                messages.error(request, "Não encontramos nenhum CPF nesse arquivo.")
            else:
                resultado = associar_cpfs(perfil, cpfs, request.user)
    else:
        form = UploadCSVForm()
    return render(
        request,
        "projetos/perfil_associar_lote.html",
        {"perfil": perfil, "form": form, "resultado": resultado},
    )


@login_required
@requer_permissao("participacoes.mover_etapa")
def perfil_associar_lote_modelo(request, pk):
    wb = Workbook()
    ws = wb.active
    ws.title = "Associação em lote"
    ws.append(["CPF"])
    celula = ws.cell(row=1, column=1)
    celula.font = Font(bold=True, color="FFFFFF")
    celula.fill = PatternFill("solid", fgColor="C4143F")
    ws.append(["111.444.777-35"])
    ws.column_dimensions["A"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="modelo_associacao_lote.xlsx"'
    return response
