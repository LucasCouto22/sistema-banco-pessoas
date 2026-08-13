"""Exportação da lista de Participações em PDF e XLSX — sempre respeitando os
mesmos filtros aplicados na tela (`?nome=&projeto=&etapa=&status=&nota=`) e a
mesma regra de mascaramento de PII usada no resto do sistema: quem não tem a
permissão `participantes.revelar_pii` recebe CPF/telefone/e-mail mascarados,
igual à tela de Banco de Pessoas."""

import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from core.relatorios import caminho_logo
from projetos.models import Projeto

from .models import Participacao

VIOLETA_DEEP = colors.HexColor("#C4143F")
VIOLETA_SOFT = colors.HexColor("#FDE8EE")
TINTA = colors.HexColor("#251A1E")
MUDO = colors.HexColor("#9A8790")
LINHA = colors.HexColor("#F1E4E8")


def _campos_participante(participante, pode_revelar_pii):
    if pode_revelar_pii:
        return participante.cpf, participante.telefone, participante.email
    return participante.cpf_mascarado, participante.telefone_mascarado, participante.email_mascarado


def _nota_formatada(avaliacao):
    if not avaliacao:
        return "Sem avaliação"
    return (
        f"{avaliacao.comunicacao}/{avaliacao.pontualidade}/{avaliacao.repertorio} "
        f"→ {avaliacao.nota_geral}"
    )


def _linhas(participacoes, pode_revelar_pii):
    linhas = []
    for part in participacoes:
        participante = part.participante
        cpf, telefone, email = _campos_participante(participante, pode_revelar_pii)
        avaliacao = getattr(part, "avaliacao", None)
        linhas.append(
            {
                "codigo": participante.codigo,
                "nome": participante.nome,
                "cpf": cpf,
                "telefone": telefone,
                "email": email or "—",
                "cidade": f"{participante.cidade}/{participante.uf}",
                "projeto": part.perfil.projeto.nome,
                "perfil": part.perfil.nome,
                "etapa": part.get_etapa_display(),
                "status": part.get_status_display() or "—",
                "responsavel": (part.responsavel.get_full_name() or part.responsavel.username)
                if part.responsavel
                else "—",
                "observacao": part.observacao or "—",
                "criado_em": timezone.localtime(part.criado_em).strftime("%d/%m/%Y %H:%M"),
                "comunicacao": avaliacao.comunicacao if avaliacao else None,
                "pontualidade": avaliacao.pontualidade if avaliacao else None,
                "repertorio": avaliacao.repertorio if avaliacao else None,
                "nota_geral": avaliacao.nota_geral if avaliacao else None,
                "nota_formatada": _nota_formatada(avaliacao),
                "comentario_avaliacao": (avaliacao.comentario if avaliacao else "") or "—",
                "avaliado_por": (
                    (avaliacao.avaliado_por.get_full_name() or avaliacao.avaliado_por.username)
                    if avaliacao and avaliacao.avaliado_por
                    else "—"
                ),
                "avaliado_em": (
                    timezone.localtime(avaliacao.avaliado_em).strftime("%d/%m/%Y %H:%M")
                    if avaliacao
                    else "—"
                ),
            }
        )
    return linhas


def _resumo_filtros(filtros):
    partes = []
    if filtros.get("nome"):
        partes.append(f"nome contém “{filtros['nome']}”")
    if filtros.get("projeto"):
        try:
            partes.append(f"projeto: {Projeto.objects.get(pk=filtros['projeto']).nome}")
        except Projeto.DoesNotExist:
            pass
    if filtros.get("etapa"):
        partes.append(f"etapa: {dict(Participacao.Etapa.choices).get(filtros['etapa'], filtros['etapa'])}")
    if filtros.get("status"):
        partes.append(f"status: {dict(Participacao.Status.choices).get(filtros['status'], filtros['status'])}")
    if filtros.get("nota") == "sem_avaliacao":
        partes.append("nota: sem avaliação")
    elif filtros.get("nota"):
        partes.append(f"nota: {filtros['nota']} ou mais")
    return "; ".join(partes) if partes else "nenhum filtro aplicado — todas as participações"


def gerar_xlsx(participacoes, pode_revelar_pii):
    linhas = _linhas(participacoes, pode_revelar_pii)

    colunas = [
        ("codigo", "Código"),
        ("nome", "Participante"),
        ("cpf", "CPF"),
        ("telefone", "Telefone"),
        ("email", "E-mail"),
        ("cidade", "Cidade/UF"),
        ("projeto", "Projeto"),
        ("perfil", "Perfil"),
        ("etapa", "Etapa"),
        ("status", "Status"),
        ("responsavel", "Responsável"),
        ("criado_em", "Entrou na participação em"),
        ("comunicacao", "Comunicação"),
        ("pontualidade", "Pontualidade"),
        ("repertorio", "Repertório"),
        ("nota_geral", "Nota geral"),
        ("comentario_avaliacao", "Comentário da avaliação"),
        ("avaliado_por", "Avaliado por"),
        ("avaliado_em", "Avaliado em"),
        ("observacao", "Observação"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Participações"

    ws.append(["Qualy Vortice — Relatório de Participações"])
    ws["A1"].font = Font(bold=True, size=14, color="C4143F")
    ws.append([f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"])
    ws["A2"].font = Font(size=9, color="9A8790")
    ws.append([])

    cabecalho_linha = 4
    ws.append([rotulo for _, rotulo in colunas])
    for col_idx in range(1, len(colunas) + 1):
        celula = ws.cell(row=cabecalho_linha, column=col_idx)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="C4143F")
        celula.alignment = Alignment(vertical="center", wrap_text=True)

    for linha in linhas:
        ws.append([linha[chave] if linha[chave] is not None else "—" for chave, _ in colunas])

    for col_idx, (chave, rotulo) in enumerate(colunas, start=1):
        maior = max([len(rotulo)] + [len(str(linha[chave])) for linha in linhas] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(maior + 2, 10), 45)

    ws.freeze_panes = f"A{cabecalho_linha + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    nome_arquivo = f"participacoes_{timezone.localtime().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def gerar_pdf(participacoes, pode_revelar_pii, filtros):
    linhas = _linhas(participacoes, pode_revelar_pii)

    estilos = getSampleStyleSheet()
    estilo_celula = ParagraphStyle(
        "celula", parent=estilos["Normal"], fontSize=7, leading=8.5, textColor=TINTA
    )
    estilo_titulo = ParagraphStyle(
        "titulo", parent=estilos["Heading1"], fontSize=16, textColor=VIOLETA_DEEP, spaceAfter=2
    )
    estilo_sub = ParagraphStyle("sub", parent=estilos["Normal"], fontSize=8.5, textColor=MUDO)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.4 * cm,
        title="Relatório de Participações — Qualy Vortice",
    )

    elementos = []

    logo = caminho_logo()
    if logo:
        cabecalho = Table(
            [[Image(logo, width=1.6 * cm, height=1.6 * cm), Paragraph("Qualy Vortice", estilo_titulo)]],
            colWidths=[1.9 * cm, None],
        )
        cabecalho.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elementos.append(cabecalho)
    else:
        elementos.append(Paragraph("Qualy Vortice", estilo_titulo))

    elementos.append(Paragraph("Relatório de Participações", estilos["Heading2"]))
    elementos.append(
        Paragraph(f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}", estilo_sub)
    )
    elementos.append(Paragraph(f"Filtros: {_resumo_filtros(filtros)}", estilo_sub))
    elementos.append(Paragraph(f"{len(linhas)} participação(ões)", estilo_sub))
    elementos.append(
        Paragraph(
            "Legenda — Nota (C/P/R → Geral): Comunicação / Pontualidade / Repertório → nota geral (média dos três)",
            estilo_sub,
        )
    )
    elementos.append(Spacer(1, 10))

    cabecalhos = [
        "Código", "Participante", "CPF", "Telefone", "E-mail", "Cidade/UF", "Projeto", "Perfil",
        "Etapa", "Status", "Responsável", "Nota (C/P/R → Geral)", "Comentário da avaliação",
        "Observação", "Entrada",
    ]
    dados = [[Paragraph(c, ParagraphStyle("cab", parent=estilo_celula, textColor=colors.white, fontName="Helvetica-Bold")) for c in cabecalhos]]
    for linha in linhas:
        dados.append(
            [
                Paragraph(linha["codigo"], estilo_celula),
                Paragraph(linha["nome"], estilo_celula),
                Paragraph(linha["cpf"], estilo_celula),
                Paragraph(linha["telefone"], estilo_celula),
                Paragraph(linha["email"], estilo_celula),
                Paragraph(linha["cidade"], estilo_celula),
                Paragraph(linha["projeto"], estilo_celula),
                Paragraph(linha["perfil"], estilo_celula),
                Paragraph(linha["etapa"], estilo_celula),
                Paragraph(linha["status"], estilo_celula),
                Paragraph(linha["responsavel"], estilo_celula),
                Paragraph(linha["nota_formatada"], estilo_celula),
                Paragraph(linha["comentario_avaliacao"], estilo_celula),
                Paragraph(linha["observacao"], estilo_celula),
                Paragraph(linha["criado_em"], estilo_celula),
            ]
        )

    largura_util = landscape(A4)[0] - 2.4 * cm
    proporcoes = [
        0.058, 0.087, 0.068, 0.058, 0.087, 0.058, 0.068, 0.058,
        0.058, 0.049, 0.068, 0.078, 0.078, 0.068, 0.058,
    ]
    larguras = [largura_util * p for p in proporcoes]

    tabela = Table(dados, colWidths=larguras, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), VIOLETA_DEEP),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, VIOLETA_SOFT]),
                ("GRID", (0, 0), (-1, -1), 0.4, LINHA),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elementos.append(tabela)

    def _rodape(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(MUDO)
        canvas.drawString(1.2 * cm, 0.8 * cm, "Qualy Vortice — documento gerado pelo sistema")
        canvas.drawRightString(
            landscape(A4)[0] - 1.2 * cm, 0.8 * cm, f"Página {_doc.page}"
        )
        canvas.restoreState()

    doc.build(elementos, onFirstPage=_rodape, onLaterPages=_rodape)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    nome_arquivo = f"participacoes_{timezone.localtime().strftime('%Y%m%d_%H%M')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response
